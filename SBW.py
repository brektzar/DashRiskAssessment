from dash import html, dcc, Input, Output, State
import dash_bootstrap_components as dbc
import pandas as pd
import openpyxl
from openpyxl.styles import Color, PatternFill, Border, Side
import copy
from io import BytesIO
from dash.exceptions import PreventUpdate

# Risk severity matrix
risk_matrix = {
    (1, 1): 1, (1, 2): 2, (1, 3): 3, (1, 4): 4,
    (2, 1): 2, (2, 2): 4, (2, 3): 6, (2, 4): 8,
    (3, 1): 3, (3, 2): 6, (3, 3): 9, (3, 4): 12,
    (4, 1): 4, (4, 2): 8, (4, 3): 12, (4, 4): 16
}

risks = []


def get_severity_info(severity):
    if 1 <= severity <= 2:
        return "Låg", "green"
    elif 3 <= severity <= 7:
        return "Medel", "yellow"
    elif 8 <= severity <= 9:
        return "Medelhög", "orange"
    else:
        return "Hög", "red"


def get_severity_color(severity):
    _, color = get_severity_info(severity)
    return {"green": "00FF00", "yellow": "FFFF00", "orange": "FFA500", "red": "FF0000"}[color]


def create_excel_file(data):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Risk Assessment"

    # Add headers
    headers = ["Riskkälla", "Riskbeskrivning", "Sannolikhet", "Konsekvens", "Allvarlighet", "Åtgärd", "Ansvarig",
               "Kommentarer", "Datum för åtgärd", "Datum för Uppföljning"]

    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)

    # Define colors for alternating rows and columns
    row_colors = ["F0F0F0", "FFFFFF"]  # Light gray and white
    col_colors = ["E6E6E6", "F0F0F0"]  # Two shades of light gray

    # Define border styles
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))
    thick_border = Border(left=Side(style='medium'),
                          right=Side(style='medium'),
                          top=Side(style='medium'),
                          bottom=Side(style='medium'))

    # Add data and format cells
    for row, item in enumerate(data, start=2):
        row_color = row_colors[(row - 2) % 2]

        for col in range(1, 11):
            cell = ws.cell(row=row, column=col)
            col_color = col_colors[(col - 1) % 2]

            # Blend row and column colors
            blended_color = blend_colors(row_color, col_color)
            cell.fill = PatternFill(start_color=blended_color, end_color=blended_color, fill_type="solid")

            # Add thin border to all cells
            cell.border = thin_border

        # Set cell values
        ws.cell(row=row, column=1, value=item["name"])
        ws.cell(row=row, column=2, value=item["description"])
        ws.cell(row=row, column=3, value=item["likelihood"])
        ws.cell(row=row, column=4, value=item["impact"])
        severity = item["severity"]
        severity_cell = ws.cell(row=row, column=5, value=severity)
        severity_cell.fill = PatternFill(start_color=get_severity_color(severity),
                                         end_color=get_severity_color(severity), fill_type="solid")
        # Add thick border to severity cell
        severity_cell.border = thick_border
        ws.cell(row=row, column=6, value=item["action"])
        ws.cell(row=row, column=7, value=item["responsible"])
        ws.cell(row=row, column=8, value=item["comments"])
        ws.cell(row=row, column=9, value=item["action_date"])
        ws.cell(row=row, column=10, value=item["follow_up_date"])

    # Add borders and thick borders to header row
    for col in range(1, 11):
        header_cell = ws.cell(row=1, column=col)
        header_cell.border = thick_border if col == 5 else thin_border

    # Set column widths
    small_columns = ['C', 'D', 'E', 'G']
    medium_columns = ['A', 'I', 'J']
    big_columns = ['B', 'F', 'H']

    for col in ws.columns:
        column_letter = col[1].column_letter
        if column_letter in small_columns:
            width = 12
        elif column_letter in medium_columns:
            width = 20
        elif column_letter in big_columns:
            width = 50
        else:
            width = None

        ws.column_dimensions[column_letter].width = width

        for cell in col:
            alignment = copy.copy(cell.alignment)
            alignment.wrapText = True
            if column_letter in ['B', 'H']:
                alignment.horizontal = 'left'
                alignment.vertical = 'center'
            else:
                alignment.horizontal = 'center'
                alignment.vertical = 'center'

            cell.alignment = alignment

    # Save to BytesIO object
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    return excel_file


def blend_colors(color1, color2):
    # Convert hex to RGB
    r1, g1, b1 = int(color1[:2], 16), int(color1[2:4], 16), int(color1[4:], 16)
    r2, g2, b2 = int(color2[:2], 16), int(color2[2:4], 16), int(color2[4:], 16)

    # Blend colors (simple average)
    r = (r1 + r2) // 2
    g = (g1 + g2) // 2
    b = (b1 + b2) // 2

    # Convert back to hex
    return f"{r:02x}{g:02x}{b:02x}"


def create_matrix_cell(value, color):
    return html.Td(value, style={
        "backgroundColor": color,
        "color": "black" if color in ["yellow", "green"] else "white",
        "border": "1px solid black",
        "padding": "10px",
        "textAlign": "center"
    })


matrix_table = html.Table([
    html.Tr([html.Th("Sannolikhet",
                     className="bg-secondary text-white",
                     colSpan=4,
                     style={"border": "1px solid black", "padding": "10px"})]),
    html.Tr([html.Th("Konsekvens",
                     className="bg-secondary text-white",
                     rowSpan=5,
                     style={"border": "1px solid black", "padding": "10px"})] + [
                html.Th(i, style={"border": "1px solid black", "padding": "10px"}) for i in range(1, 5)]),
    html.Tr([html.Th("4", style={"border": "1px solid black", "padding": "10px"})] + [
        create_matrix_cell(risk_matrix[(4, i)], get_severity_info(risk_matrix[(4, i)])[1]) for i in range(1, 5)]),
    html.Tr([html.Th("3", style={"border": "1px solid black", "padding": "10px"})] + [
        create_matrix_cell(risk_matrix[(3, i)], get_severity_info(risk_matrix[(3, i)])[1]) for i in range(1, 5)]),
    html.Tr([html.Th("2", style={"border": "1px solid black", "padding": "10px"})] + [
        create_matrix_cell(risk_matrix[(2, i)], get_severity_info(risk_matrix[(2, i)])[1]) for i in range(1, 5)]),
    html.Tr([html.Th("1", style={"border": "1px solid black", "padding": "10px"})] + [
        create_matrix_cell(risk_matrix[(1, i)], get_severity_info(risk_matrix[(1, i)])[1]) for i in range(1, 5)]),
], style={"borderCollapse": "collapse", "marginBottom": "20px"})

severity_descriptions = html.Div([
    html.H3("Allvarlighetsgrad"),
    html.Div([
        html.Span("Låg",
                  style={"backgroundColor": "green", "color": "black", "padding": "2px 5px", "borderRadius": "3px",
                         "marginRight": "10px"}),
        html.Span("Medel",
                  style={"backgroundColor": "yellow", "color": "black", "padding": "2px 5px", "borderRadius": "3px",
                         "marginRight": "10px"}),
        html.Span("Medelhög",
                  style={"backgroundColor": "orange", "color": "white", "padding": "2px 5px", "borderRadius": "3px",
                         "marginRight": "10px"}),
        html.Span("Hög",
                  style={"backgroundColor": "red", "color": "white", "padding": "2px 5px", "borderRadius": "3px"}),
    ])
])

# Define the layout
layout = html.Div([
    html.Div([
        html.P("Riskbedömning för arbete",
               className="bg-primary text-white py-3 fs-3 fw-bolder text-center"),
    ], style={"marginTop": "20px", "marginBottom": "30px"}),
    html.Div([
        html.Div([
            html.Div([
                html.Div([
                    html.P("Skriv in vart risken har uppstått",
                           className="bg-primary text-white py-1 rounded-pill fw-bold text-center"),
                ], style={"marginTop": "20px"}),
                dcc.Input(id="risk-name", type="text", placeholder="Riskkälla",
                          style={"marginRight": "10px", "width": "200px"}),
            ], style={"marginBottom": "15px"}),

            html.Div([
                html.Div([
                    html.P("Beskriv risken tydligt",
                           className="bg-primary text-white py-1 rounded-pill fw-bold text-center"),
                ], style={"marginTop": "20px"}),
                dcc.Textarea(id="risk-description",
                             placeholder="Riskbeskrivning",
                             style={"width": "calc(100% - 210px)", "height": "100px"}),
            ], style={"marginBottom": "15px"}),

            html.Div([
                html.Div([
                    html.P("Skatta risken utefter sannolikhet och konsekvens",
                           className="bg-primary text-white py-1 rounded-pill fw-bold text-center"),
                ], style={"marginTop": "20px"}),
                dcc.Dropdown(id="likelihood", options=[{"label": i, "value": i} for i in range(1, 5)],
                             placeholder="Sannolikhet",
                             style={"width": "200px", "marginRight": "10px", "display": "inline-block"}),
                dcc.Dropdown(id="impact", options=[{"label": i, "value": i} for i in range(1, 5)],
                             placeholder="Konsekvens",
                             style={"width": "200px", "display": "inline-block"}),
            ], style={"marginBottom": "15px"}),

            html.Div([
                html.Div([
                    html.P("Skriv ner åtgärd för att hantera risken",
                           className="bg-primary text-white py-1 rounded-pill fw-bold text-center"),
                ], style={"marginTop": "20px"}),
                dcc.Input(id="action", type="text", placeholder="Åtgärd",
                          style={"marginRight": "10px", "width": "300px"}),
            ], style={"marginBottom": "15px"}),

            html.Div([
                html.Div([  # ---------------------------------------------------------------------------------#
                    html.Div([
                        html.P("Välj datum för åtgärd",
                               className="bg-primary text-white py-1 rounded-pill fw-bold text-center"),
                    ], style={"marginTop": "20px"}),
                    dcc.DatePickerSingle(
                        id="action-date",
                        placeholder="Välj datum",
                        style={"width": "calc(50%)"}
                    ),
                ], style={"marginBottom": "20px"}),

                html.Div([  # ---------------------------------------------------------------------------------#
                    html.Div([
                        html.P("Välj datum för uppföljning",
                               className="bg-primary text-white py-1 rounded-pill fw-bold text-center"),
                    ], style={"marginTop": "20px"}),
                    dcc.DatePickerSingle(
                        id="follow-up-date",
                        placeholder="Välj datum",
                        style={"width": "calc(50% - 5px)", 'font-size': '6px', 'display': 'inline-block'}
                    ),
                ], style={"marginBottom": "20px"}),
            ], style={"marginBottom": "20px"}),

            html.Div([
                html.Div([
                    html.Div([
                        html.P("Välj den person som ansvarar för att åtgärden blir fullföljd",
                               className="bg-primary text-white py-1 rounded-pill fw-bold text-center"),
                    ], style={"marginTop": "20px"}),
                    dcc.Input(id="responsible", type="text", placeholder="Ansvarig",
                              style={"marginRight": "calc(100% - 10px)", "width": "200px", "display": "flex"}),
                ], style={"marginBottom": "15px"}),

                html.Div([
                    html.P("Lägg till eventuella kommentarer om risken (valfritt)",
                           className="bg-primary text-white py-1 rounded-pill fw-bold text-center"),
                ], style={"marginTop": "20px"}),
                dcc.Textarea(id="comments", placeholder="Kommentarer",
                             style={"width": "calc(100% - 210px)", "height": "100px"}),
            ], style={"marginBottom": "15px"}),

            html.Div([
                html.P("Lägg till risk eller skapa Excel-fil",
                       className="bg-primary text-white py-1 rounded-pill fw-bold text-center"),
            ], style={"marginTop": "20px"}),
            dbc.Button("Lägg till risk", id="add-risk-button", color="info", className="me-1"),
            dbc.Button("Generera Excel", id="generate-excel-button", color="success", className="me-1"),
            dcc.Download(id="download-excel"),
            html.Div(id="risks-list", style={"marginBottom": "20px"}),
        ], style={"width": "60%", "float": "left"}),

        html.Div([
            matrix_table,
            severity_descriptions
        ], style={"width": "35%", "float": "right"})
    ], style={"overflow": "hidden"}),
], style={"border": "1px solid #ccc", "maxWidth": "1200px", "height": "auto", "margin": "0 auto", "padding": "20px"})


# Define the callbacks
def register_callbacks(app):
    @app.callback(
        Output("risks-list", "children"),
        Input("add-risk-button", "n_clicks"),
        State("risk-name", "value"),
        State("risk-description", "value"),
        State("likelihood", "value"),
        State("impact", "value"),
        State("action", "value"),
        State("responsible", "value"),
        State("comments", "value"),
        State("action-date", "date"),
        State("follow-up-date", "date"),
        prevent_initial_call=True
    )
    def add_risk(n_clicks, name, description, likelihood, impact, action, responsible, comments, action_date,
                 follow_up_date):
        if not all([name, description, likelihood, impact, action, responsible, action_date,
                    follow_up_date]):
            raise PreventUpdate

        severity = risk_matrix[(likelihood, impact)]
        severity_label, severity_color = get_severity_info(severity)

        risks.append({
            "name": name,
            "description": description,
            "likelihood": likelihood,
            "impact": impact,
            "severity": severity,
            "severity_label": severity_label,
            "severity_color": severity_color,
            "action": action,
            "responsible": responsible,
            "comments": comments,
            "action_date": action_date,
            "follow_up_date": follow_up_date
        })

        return html.Ul([
            html.Li([
                f"{risk['name']} - "
                f"{risk['description']} - "
                f"Sannolikhet: {risk['likelihood']}, "
                f"Konsekvens: {risk['impact']} ",
                html.Span(
                    risk['severity_label'],
                    style={
                        "backgroundColor": risk['severity_color'],
                        "color": "black" if risk['severity_color'] in ["yellow", "green"] else "white",
                        "padding": "2px 5px",
                        "borderRadius": "3px",
                        "marginLeft": "5px"
                    }
                ),
                html.Br(),
                f"Åtgärd: {risk['action']}",
                html.Br(),
                f"Ansvarig: {risk['responsible']}",
                html.Br(),
                f"Kommentarer: {risk['comments']}" if risk['comments'] else "",
                html.Br(),
                f"Datum för åtgärd: {risk['action_date']}" if risk['action_date'] else "",
                html.Br(),
                f"Datum för Uppföljning: {risk['follow_up_date']}" if risk['follow_up_date'] else ""
            ]) for risk in risks
        ])

    @app.callback(
        Output("download-excel", "data"),
        Input("generate-excel-button", "n_clicks"),
        prevent_initial_call=True
    )
    def generate_excel(_):
        if not risks:
            raise PreventUpdate

        excel_file = create_excel_file(risks)
        return dcc.send_bytes(excel_file.getvalue(), "risk_assessment.xlsx")
